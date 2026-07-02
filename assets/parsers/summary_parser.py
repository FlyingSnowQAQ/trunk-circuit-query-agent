"""
汇总Excel解析器

解析"1-移动干线电路名称汇总.xlsx"的15个sheet，
按列映射规则统一入库。
"""
import json
import openpyxl
from .column_mapper import SHEET_MAP


def parse_summary_excel(file_path: str) -> list:
    """
    解析汇总Excel，返回记录列表
    每一条记录是一个dict，字段名与circuits表一致
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rules = SHEET_MAP.get(sheet_name, None)
        if rules is None:
            # 未匹配的sheet，尝试按默认方式解析
            rules = _guess_rules(sheet_name, ws)

        phase = rules["phase"]
        header_rows = rules["header_rows"]
        extra_skip = rules["extra_skip_rows"]
        skip_empty = rules["skip_empty_circuit"]
        col_map = rules["columns"]

        # 收集所有行
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_rows:
            continue

        # 表头行
        header = rows[header_rows]
        # 构建列名 -> 列索引 映射（使用字符串匹配）
        col_index = {}
        for i, h in enumerate(header):
            if h is not None:
                h_str = str(h).strip()
                if h_str in col_map:
                    col_index[i] = col_map[h_str]

        # 数据行
        data_start = header_rows + 1
        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            # 跳过全空行
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            record = {
                "project_phase": phase,
                "source_file": file_path,
                "source_sheet": sheet_name,
            }

            has_circuit_no = False
            raw_data = {}

            for col_idx, target_field in col_index.items():
                if col_idx < len(row):
                    val = row[col_idx]
                else:
                    val = None

                # 转为字符串
                if val is not None:
                    val_str = str(val).strip()
                else:
                    val_str = ""

                raw_data[f"col_{col_idx}"] = val_str

                # 忽略映射为None的字段
                if target_field is None:
                    continue

                if target_field == "circuit_no":
                    if val_str:
                        has_circuit_no = True
                    # 兼容"10期"等使用"电路号"列名的情况
                    record[target_field] = val_str

                elif target_field in (
                    "system_name", "circuit_name", "endpoint",
                    "system_type", "capacity", "business_nature",
                    "service_type", "protection_type", "route_nature",
                    "protection_circuit", "protection_endpoint",
                    "remark", "update_date", "system_category"
                ):
                    # 特殊处理：11.1期"接口种类"被映射到capacity
                    record[target_field] = val_str

            # 对于13期及以后（有跳次/路由信息的），电路摘要仅存基本字段
            # 路由详情由route_parser通过路由表文件处理

            # 跳过空记录
            if skip_empty and not has_circuit_no:
                continue

            record["raw_json"] = json.dumps(raw_data, ensure_ascii=False)
            all_records.append(record)

    wb.close()
    return all_records


def _guess_rules(sheet_name: str, ws) -> dict:
    """
    对于未在SHEET_MAP中配置的sheet，自动推测规则
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"phase": sheet_name, "header_rows": 0, "extra_skip_rows": 0,
                "skip_empty_circuit": True, "columns": {}}

    # 尝试在第一行找常见列名
    header = rows[0]
    col_map = {}
    known_columns = {
        "电路编号": "circuit_no",
        "电路号": "circuit_no",
        "系统名称": "system_name",
        "系统": "system_name",
        "电路名称": "circuit_name",
        "业务名称": "circuit_name",
        "终端点": "endpoint",
        "建设期": "project_phase",
        "传输建设期": "project_phase",
        "系统性质": "system_type",
        "容量": "capacity",
        "业务性质": "business_nature",
        "保护类型": "protection_type",
        "业务种类": "service_type",
    }
    for i, h in enumerate(header):
        if h is not None:
            h_str = str(h).strip()
            if h_str in known_columns:
                col_map[h_str] = known_columns[h_str]

    return {
        "phase": sheet_name,
        "header_rows": 0,
        "extra_skip_rows": 0,
        "skip_empty_circuit": True,
        "columns": col_map
    }

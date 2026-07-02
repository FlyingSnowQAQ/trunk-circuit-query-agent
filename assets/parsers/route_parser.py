"""
路由表统一解析器

根据建设期和文件类型，将不同格式的路由表文件解析为统一结构。
数据源: 2-光通道路由表 目录下的各期文件
"""
import json
import os
import openpyxl


def parse_routing_table(file_path: str, phase: str, file_type: str) -> list:
    """
    解析单一路由表文件，返回记录列表
    
    参数:
        file_path: 文件路径
        phase: 建设期标识
        file_type: "xlsx", "xls", "doc"
    
    返回: list of dict，每个dict为一条记录
    """
    if file_type == "xlsx":
        return _parse_xlsx_routing(file_path, phase)
    elif file_type == "xls":
        return _parse_xls_routing(file_path, phase)
    elif file_type == "doc":
        return _parse_doc_routing(file_path, phase)
    else:
        raise ValueError(f"不支持的文件类型: {file_type}")


def _parse_xlsx_routing(file_path: str, phase: str) -> list:
    """解析 .xlsx 路由表"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 查找表头行: 找包含"电路号"/"系统名称"/"电路名称"/"系统"等关键字的行
        header_row_idx = _find_header_row(rows)
        if header_row_idx is None:
            continue

        header = rows[header_row_idx]
        header_str = [str(h).strip() if h is not None else "" for h in header]

        # 识别列类型
        col_types = _identify_xlsx_columns(header_str, phase)

        # 解析数据行
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            record = _extract_record_from_row(row, col_types, phase, file_path, sheet_name)
            if record:
                all_records.append(record)

    wb.close()
    return all_records


def _parse_xls_routing(file_path: str, phase: str) -> list:
    """解析 .xls 路由表"""
    from .xls_parser import parse_xls_summary
    records = parse_xls_summary(file_path)

    all_records = []
    # 将平坦数据按sheet分组处理
    from collections import defaultdict
    sheet_groups = defaultdict(list)
    for r in records:
        sheet_groups[r["sheet"]].append(r)

    for sheet_name, group in sheet_groups.items():
        rows = [r["row_data"] for r in group]

        header_row_idx = _find_header_row(rows)
        if header_row_idx is None:
            continue

        header = rows[header_row_idx]
        header_str = [str(h).strip() if h is not None else "" for h in header]

        col_types = _identify_xlsx_columns(header_str, phase)

        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            record = _extract_record_from_row(row, col_types, phase, file_path, sheet_name)
            if record:
                all_records.append(record)

    return all_records


def _parse_doc_routing(file_path: str, phase: str) -> list:
    """解析 .doc 路由表"""
    from .doc_parser import parse_doc
    try:
        rows = parse_doc(file_path)
    except RuntimeError as e:
        # .doc 解析失败，记录警告
        return [{"_error": str(e), "source_file": file_path, "project_phase": phase}]

    all_records = []
    header = rows[0] if rows else []
    header_str = [str(h).strip() for h in header]
    col_types = _identify_xlsx_columns(header_str, phase)

    for row_idx in range(1, len(rows)):
        row = rows[row_idx]
        if isinstance(row, dict) and row.get("is_paragraph"):
            continue
        record = _extract_record_from_row(row, col_types, phase, file_path, "7期")
        if record:
            all_records.append(record)

    return all_records


# ========== 列类型识别 ==========

COLUMN_PATTERNS = {
    "circuit_no": ["电路号", "电路编号"],
    "system_name": ["系统名称", "系统", "系统名"],
    "circuit_name": ["电路名称", "业务名称", "电路名"],
    "endpoint": ["终端点", "终端"],
    "station_a": ["A站", "A站点", "A端站", "路由A站", "A-站点", "A端局站"],
    "station_b": ["B站", "B站点", "B端站", "路由B站", "B-站点", "B端局站"],
    "timeslot_id": ["时隙ID", "时隙", "时隙号", "光通路时隙"],
    "multiplex_section": ["复用段名称", "复用段", "段名称"],
    "device_type": ["设备类型", "设备型号"],
    "hop_order": ["跳次", "序号", "行号", "链路ID"],
    "route_type": ["主备", "路由性质", "主用/备用"],
    "province": ["省份"],
    "station": ["局站", "站点", "局_站"],
    "platform": ["平台"],
    "protection": ["保护方式", "保护类型"],
    "rate": ["速率", "容量"],
    "link_id": ["链路ID", "链路编号"],
    "terminal_type": ["终端/转接", "终端转接"],
    "usage_type": ["本期使用", "本期"],
    "usage_nature": ["使用性质"],
    "a_equipment_id": ["A-网元ID", "A网元ID", "A网元"],
    "a_line_slot": ["A-线路槽位", "A线路槽位"],
    "a_line_port": ["A-线路端口", "A线路端口"],
    "a_line_board": ["A-线路板类型", "A线路板"],
    "a_tributary_slot": ["A-支路槽位", "A支路槽位"],
    "a_tributary_port": ["A-支路端口", "A支路端口"],
    "a_tributary_board": ["A-支路板类型", "A支路板"],
    "b_equipment_id": ["B-网元ID", "B网元ID", "B网元"],
    "b_line_slot": ["B-线路槽位", "B线路槽位"],
    "b_line_port": ["B-线路端口", "B线路端口"],
    "b_line_board": ["B-线路板类型", "B线路板"],
    "b_tributary_slot": ["B-支路槽位", "B支路槽位"],
    "b_tributary_port": ["B-支路端口", "B支路端口"],
    "b_tributary_board": ["B-支路板类型", "B支路板"],
    "rack_code": ["机架编码", "机架号"],
    "slot_port": ["槽位端口", "槽位"],
    "timeslot": ["时隙"],
    "config_phase": ["配置期", "配置阶段"],
    "use_phase": ["使用期", "使用阶段"],
    "multiplex_section_name": ["复用段名称"],
}


def _find_header_row(rows: list) -> int:
    """在行列表中找到表头行索引"""
    keywords = {"电路号", "电路编号", "系统名称", "系统", "电路名称",
                "终端点", "A站", "B站", "时隙", "链路ID", "省份"}
    for idx, row in enumerate(rows):
        if idx > 20:  # 最多找前20行
            break
        if row:
            row_strs = [str(v).strip() for v in row if v is not None]
            matched = sum(1 for k in keywords for r in row_strs if k in r)
            if matched >= 2:
                return idx
    return None


def _identify_xlsx_columns(header: list, phase: str) -> dict:
    """
    识别每列的数据类型
    返回: { col_index: "field_name" }
    """
    col_types = {}
    for idx, h in enumerate(header):
        for field_name, patterns in COLUMN_PATTERNS.items():
            if any(p in h for p in patterns):
                col_types[idx] = field_name
                break
    return col_types


def _extract_record_from_row(row: tuple, col_types: dict, phase: str,
                              file_path: str, sheet_name: str) -> dict:
    """从单行数据中提取记录"""
    # 先判断记录类型
    has_hop_fields = any(f in col_types.values() for f in
                         ["station_a", "station_b", "hop_order", "timeslot_id"])
    has_mux_fields = any(f in col_types.values() for f in
                         ["province", "station", "multiplex_section_name", "link_id"])
    has_device_fields = any(f in col_types.values() for f in
                            ["a_equipment_id", "a_line_slot", "b_equipment_id"])

    raw_data = {}
    record = {
        "project_phase": phase,
        "source_file": file_path,
        "source_sheet": sheet_name,
    }

    for col_idx, field_name in col_types.items():
        if col_idx >= len(row):
            continue
        val = row[col_idx]
        val_str = str(val).strip() if val is not None else ""
        raw_data[f"col_{col_idx}_{field_name}"] = val_str
        record[field_name] = val_str

    record["raw_json"] = json.dumps(raw_data, ensure_ascii=False)

    # 判断记录应存入哪个表
    if has_hop_fields or has_device_fields:
        record["_table"] = "circuit_hops"
    elif has_mux_fields:
        record["_table"] = "multiplex_sections"
    else:
        record["_table"] = "circuits"

    return record


def batch_parse_routing_dir(route_dir: str) -> dict:
    """
    批量解析路由表目录中的所有文件
    返回: { phase: [records] }
    
    此函数从 column_mapper 获取文件映射，自动检测文件格式并解析
    """
    from .column_mapper import ROUTE_FILE_MAP

    # 列出目录中的所有文件
    if not os.path.isdir(route_dir):
        return {"_error": f"目录不存在: {route_dir}"}

    files = [f for f in os.listdir(route_dir)
             if os.path.isfile(os.path.join(route_dir, f))
             and not f.startswith("~$")]  # 跳过临时文件

    # 文件 -> 建设期 映射
    file_phase_map = {}
    for keyword, phase, ftype, _, _ in ROUTE_FILE_MAP:
        for fname in files:
            if keyword in fname and fname.lower().endswith(('.xlsx', '.xls', '.doc', '.docx')):
                fpath = os.path.join(route_dir, fname)
                if keyword not in file_phase_map or fname.find(keyword) >= 0:
                    file_phase_map.setdefault(phase, []).append({
                        "path": fpath,
                        "type": ftype,
                    })

    # 逐文件解析
    results = {}
    for phase, entries in file_phase_map.items():
        phase_records = []
        for entry in entries:
            try:
                records = parse_routing_table(
                    entry["path"], phase, entry["type"]
                )
                phase_records.extend(records)
            except Exception as e:
                phase_records.append({
                    "_error": str(e),
                    "source_file": entry["path"],
                    "project_phase": phase,
                })
        results[phase] = phase_records

    return results
